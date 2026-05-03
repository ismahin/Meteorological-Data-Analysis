from __future__ import annotations

import argparse
import calendar
import re
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = PROJECT_ROOT / "Four Years data (2021 to 2024) 4 paramters.xlsx"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "data" / "processed" / "bmd_stations_3hourly"
DEFAULT_REPORT = PROJECT_ROOT / "outputs" / "tables" / "bmd_split_report.csv"

START_YEAR = 2021
END_YEAR = 2024
OBSERVED_HOURS = (0, 3, 6, 9, 12, 15, 18, 21)
OUTPUT_COLUMNS = ["YEAR", "MO", "DY", "HR", "T2M", "RH2M", "PRECTOTCORR", "WS10M"]

STANDARD_SHEETS = {
    "Dry bulb Temp": "T2M",
    "RH DATA": "RH2M",
    "Rainfall": "PRECTOTCORR",
}
WIND_SHEET = "Wind"
WIND_VARIABLE = "WS10M"


@dataclass
class StationBlock:
    sheet: str
    variable: str
    station_name: str
    marker_row: int
    station_index: int | None = None
    rows: list[dict[str, Any]] = field(default_factory=list)
    invalid_dates: int = 0
    skipped_hours: int = 0


def station_name_from_cell(value: str) -> str:
    match = re.search(r"Station\s*:\s*(.*?)(?:Lat\.|$)", value)
    if not match:
        return value.strip()
    return re.sub(r"\s+", " ", match.group(1).strip())


def safe_station_id(station_name: str) -> str:
    value = station_name.lower().replace("&", "and")
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_")


def is_station_marker(value: Any) -> bool:
    return isinstance(value, str) and "Station :" in value


def as_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        stripped = value.strip()
        if re.fullmatch(r"\d+", stripped):
            return int(stripped)
    return None


def valid_observation_date(year: int, month: int, day: int, hour: int) -> bool:
    if year < START_YEAR or year > END_YEAR:
        return False
    if month < 1 or month > 12:
        return False
    if hour not in OBSERVED_HOURS:
        return False
    return 1 <= day <= calendar.monthrange(year, month)[1]


def clean_value(value: Any) -> Any:
    if value is None:
        return pd.NA
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or stripped in {"-", "--", "NA", "N/A", "na", "n/a"}:
            return pd.NA
        try:
            return float(stripped)
        except ValueError:
            return pd.NA
    return value


def parse_standard_sheet(workbook_path: Path, sheet_name: str, variable: str) -> list[StationBlock]:
    ws = load_workbook(workbook_path, read_only=True, data_only=True)[sheet_name]
    blocks: list[StationBlock] = []
    current: StationBlock | None = None

    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        first_cell = row[0] if row else None
        if is_station_marker(first_cell):
            current = StationBlock(
                sheet=sheet_name,
                variable=variable,
                station_name=station_name_from_cell(first_cell),
                marker_row=row_number,
            )
            blocks.append(current)
            continue

        if current is None:
            continue

        station_index = as_int(row[0] if len(row) > 0 else None)
        year = as_int(row[1] if len(row) > 1 else None)
        month = as_int(row[2] if len(row) > 2 else None)
        hour = as_int(row[3] if len(row) > 3 else None)
        if station_index is None or year is None or month is None or hour is None:
            continue

        if current.station_index is None:
            current.station_index = station_index
        if hour not in OBSERVED_HOURS:
            current.skipped_hours += 1
            continue

        for day in range(1, 32):
            if not valid_observation_date(year, month, day, hour):
                current.invalid_dates += 1
                continue
            current.rows.append(
                {
                    "YEAR": year,
                    "MO": month,
                    "DY": day,
                    "HR": hour,
                    variable: clean_value(row[day + 3] if len(row) > day + 3 else None),
                }
            )

    return blocks


def parse_wind_speed(field_value: str) -> Any:
    field_value = field_value.strip()
    if not field_value:
        return pd.NA
    tokens = field_value.split()
    if len(tokens) > 1:
        compact = "".join(tokens)
    else:
        compact = re.sub(r"\D", "", field_value)
    if not compact:
        return pd.NA
    if len(compact) == 1:
        speed = compact
    elif len(compact) == 2:
        speed = compact[0]
    else:
        speed = compact[:-2]
    try:
        return float(speed)
    except ValueError:
        return pd.NA


def parse_wind_speed_fields(line: str) -> list[Any]:
    tokens = line.split()[4:]
    speeds: list[Any] = []
    index = 0

    while index < len(tokens) and len(speeds) < 31:
        token = tokens[index]
        if re.fullmatch(r"\d{3,4}", token):
            speeds.append(parse_wind_speed(token))
            index += 1
            continue

        if re.fullmatch(r"\d{1,2}", token):
            field_tokens = [token]
            if index + 1 < len(tokens) and re.fullmatch(r"\d{1,2}", tokens[index + 1]):
                field_tokens.append(tokens[index + 1])
                index += 2
            else:
                index += 1
            speeds.append(parse_wind_speed(" ".join(field_tokens)))
            continue

        speeds.append(pd.NA)
        index += 1

    if len(speeds) < 31:
        speeds.extend([pd.NA] * (31 - len(speeds)))
    return speeds[:31]


def parse_wind_sheet(workbook_path: Path) -> list[StationBlock]:
    ws = load_workbook(workbook_path, read_only=True, data_only=True)[WIND_SHEET]
    blocks: list[StationBlock] = []
    current: StationBlock | None = None

    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        value = row[0] if row else None
        if is_station_marker(value):
            current = StationBlock(
                sheet=WIND_SHEET,
                variable=WIND_VARIABLE,
                station_name=station_name_from_cell(value),
                marker_row=row_number,
            )
            blocks.append(current)
            continue

        if not isinstance(value, str):
            continue

        if current is None:
            continue
        if not re.match(r"^\s*\d+\s+\d{4}\s+\d+\s+\d+", value):
            continue

        meta = value.split()[:4]
        if len(meta) < 4:
            continue

        station_index = as_int(meta[0])
        year = as_int(meta[1])
        month = as_int(meta[2])
        hour = as_int(meta[3])
        if station_index is None or year is None or month is None or hour is None:
            continue

        if current.station_index is None:
            current.station_index = station_index
        if hour not in OBSERVED_HOURS:
            current.skipped_hours += 1
            continue

        speeds = parse_wind_speed_fields(value)
        for day in range(1, 32):
            if not valid_observation_date(year, month, day, hour):
                current.invalid_dates += 1
                continue
            current.rows.append(
                {
                    "YEAR": year,
                    "MO": month,
                    "DY": day,
                    "HR": hour,
                    WIND_VARIABLE: speeds[day - 1],
                }
            )

    return blocks


def select_blocks(blocks_by_variable: dict[str, list[StationBlock]]) -> tuple[dict[str, dict[str, StationBlock]], list[dict[str, Any]]]:
    station_sets = {
        variable: {block.station_name for block in blocks}
        for variable, blocks in blocks_by_variable.items()
    }
    common_stations = set.intersection(*station_sets.values())
    selected: dict[str, dict[str, StationBlock]] = {station: {} for station in common_stations}
    skipped_duplicates: list[dict[str, Any]] = []

    for variable, blocks in blocks_by_variable.items():
        by_station: dict[str, list[StationBlock]] = {}
        for block in blocks:
            by_station.setdefault(block.station_name, []).append(block)

        for station_name in common_stations:
            candidates = by_station[station_name]
            chosen = candidates[0]
            if variable == "PRECTOTCORR" and station_name == "Hatiya":
                matching = [block for block in candidates if block.station_index == 11814]
                if matching:
                    chosen = matching[0]
            for block in candidates:
                if block is not chosen:
                    skipped_duplicates.append(
                        {
                            "station_name": block.station_name,
                            "variable": block.variable,
                            "sheet": block.sheet,
                            "marker_row": block.marker_row,
                            "station_index": block.station_index,
                            "reason": "duplicate_station_block_not_selected",
                        }
                    )
            selected[station_name][variable] = chosen

    return selected, skipped_duplicates


def block_to_dataframe(block: StationBlock) -> pd.DataFrame:
    df = pd.DataFrame(block.rows)
    if df.empty:
        return pd.DataFrame(columns=["YEAR", "MO", "DY", "HR", block.variable])
    df = df.drop_duplicates(subset=["YEAR", "MO", "DY", "HR"], keep="first")
    return df[["YEAR", "MO", "DY", "HR", block.variable]]


def base_timeline() -> pd.DataFrame:
    timestamps = pd.date_range(
        f"{START_YEAR}-01-01 00:00:00",
        f"{END_YEAR}-12-31 21:00:00",
        freq="3h",
    )
    return pd.DataFrame(
        {
            "YEAR": timestamps.year,
            "MO": timestamps.month,
            "DY": timestamps.day,
            "HR": timestamps.hour,
        }
    )


def summarize_detection(blocks_by_variable: dict[str, list[StationBlock]]) -> None:
    print("Detected station blocks:")
    for variable, blocks in blocks_by_variable.items():
        counts = Counter(block.station_name for block in blocks)
        duplicates = {name: count for name, count in counts.items() if count > 1}
        print(
            f"  {variable}: {len(blocks)} blocks, "
            f"{len(counts)} unique stations"
            + (f", duplicates={duplicates}" if duplicates else "")
        )
    station_sets = [
        {block.station_name for block in blocks}
        for blocks in blocks_by_variable.values()
    ]
    common_stations = sorted(set.intersection(*station_sets))
    print(f"Common stations across all four variables: {len(common_stations)}")
    print(", ".join(common_stations))


def convert_workbook(workbook_path: Path, output_dir: Path, report_path: Path, dry_run: bool) -> None:
    blocks_by_variable: dict[str, list[StationBlock]] = {}
    for sheet_name, variable in STANDARD_SHEETS.items():
        blocks_by_variable[variable] = parse_standard_sheet(workbook_path, sheet_name, variable)
    blocks_by_variable[WIND_VARIABLE] = parse_wind_sheet(workbook_path)

    summarize_detection(blocks_by_variable)
    selected, skipped_duplicates = select_blocks(blocks_by_variable)

    if dry_run:
        return

    output_dir.mkdir(parents=True, exist_ok=True)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    timeline = base_timeline()
    report_rows: list[dict[str, Any]] = []

    for station_name in sorted(selected):
        station_df = timeline.copy()
        variable_blocks = selected[station_name]

        for variable in ["T2M", "RH2M", "PRECTOTCORR", "WS10M"]:
            station_df = station_df.merge(
                block_to_dataframe(variable_blocks[variable]),
                on=["YEAR", "MO", "DY", "HR"],
                how="left",
            )

        station_df = station_df[OUTPUT_COLUMNS]
        output_path = output_dir / f"{safe_station_id(station_name)}.csv"
        station_df.to_csv(output_path, index=False)

        row: dict[str, Any] = {
            "station_name": station_name,
            "station_id": safe_station_id(station_name),
            "output_file": str(output_path.relative_to(PROJECT_ROOT)),
            "row_count": len(station_df),
        }
        for variable in ["T2M", "RH2M", "PRECTOTCORR", "WS10M"]:
            block = variable_blocks[variable]
            row[f"{variable}_missing_count"] = int(station_df[variable].isna().sum())
            row[f"{variable}_source_sheet"] = block.sheet
            row[f"{variable}_station_index"] = block.station_index
            row[f"{variable}_source_rows"] = len(block.rows)
            row[f"{variable}_invalid_dates_skipped"] = block.invalid_dates
        report_rows.append(row)

    duplicate_rows = [
        {
            "station_name": item["station_name"],
            "station_id": safe_station_id(item["station_name"]),
            "output_file": "",
            "row_count": "",
            "note": (
                f"Skipped duplicate {item['variable']} block from {item['sheet']} "
                f"row {item['marker_row']} with station index {item['station_index']}"
            ),
        }
        for item in skipped_duplicates
    ]
    report = pd.DataFrame(report_rows + duplicate_rows)
    report.to_csv(report_path, index=False)
    print(f"Wrote {len(selected)} station CSV files to {output_dir}")
    print(f"Wrote conversion report to {report_path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Split the BMD four-parameter workbook into one 3-hourly CSV per station."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not args.workbook.exists():
        raise FileNotFoundError(args.workbook)
    convert_workbook(
        workbook_path=args.workbook,
        output_dir=args.output_dir,
        report_path=args.report,
        dry_run=args.dry_run,
    )


if __name__ == "__main__":
    main()
